import openpyxl as xl
from docxtpl import DocxTemplate


def format_cell_value(cell_value: int | float) -> str:
    if type(cell_value) == int:                      #Добавляет nuli, если cell целое число
        return f'{cell_value},0'
    else:
        return str(cell_value).replace('.', ',')


def make_context_values(sheet: xl.worksheet) -> list:
    """Делает список словарей для таблиц по ключам и номерам столбцов из table_cells для листа sheet"""
    table = []
    num = 1
    for i in range(3, 33):
        data = {}
        if sheet.cell(row=i, column=21).value == '+':
            data['num'] = num
            data['name'] = sheet.cell(i, 4).value
            data['ed1'] = sheet.cell(i, 5).value
            data['ed2'] = sheet.cell(i, 6).value
            data['do'] = format_cell_value(sheet.cell(i, 7).value)
            data['posle'] = format_cell_value(sheet.cell(i, 8).value)
            data['gesn'] = sheet.cell(i, 10).value
            table.append(data)
            num += 1
            print(f'обработка строки {num}')
    return table


def make_context_main(sheet: xl.worksheet) -> dict:
    print(f'обработка листа {sheet}')
    context = {
        'num': sheet['C2'].value,
        'name': sheet['D2'].value,
        'raboty': make_context_values(sheet),
    }
    return context


def collect_contexts(sheet_names: list) -> list:
    return [make_context_main(sheet) for sheet in sheet_names]

def return_sheets(workbook: xl.Workbook) -> list[xl.worksheet]:
    return [sheet for sheet in workbook.worksheets if sheet.title not in ['Лист2', 'Ед. расценки', 'Исх', 'Ведомость ', 'за 1 км']]


def main(file: str, template_name: str, output_file: str) -> None:
    workbook = xl.load_workbook(file, data_only=True)
    sheet_names = return_sheets(workbook)
    context = {'items':collect_contexts(sheet_names)}
    print(type(context))
    template = DocxTemplate(template_name)
    template.render(context)
    template.save(output_file)


if __name__ == '__main__':
    # workbook = xl.load_workbook('Объемы.xlsx', data_only=True)
    # print([sheet.title for sheet in workbook.worksheets])
    main('Объемы.xlsx', 'Шаблон объемы.docx', 'Ведомость объемов.docx')



