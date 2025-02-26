import openpyxl as xl

from services import format_int_value, format_str_or_digit_value

workbook = xl.load_workbook('Ведомость тест.xlsx', data_only=True)
sheet_names = [i for i in workbook.sheetnames if i not in ['Лист1', 'ИД', 'В обсл', 'аб1']]
sheet = workbook['У 1']


def context_table(table_cells, sheet):
    """Делает список словарей для таблиц по ключам и номерам столбцов из table_cells для листа sheet"""
    table = []
    for i in range(12, len(sheet['A'])):
        if sheet.cell(row=i, column=1).value not in [None, 'None']:
            table.append({key: sheet.cell(i, table_cells[key]).value for key in table_cells})
    # print(table)
    return table


table_2_cells = {'km_nach': 45, 'km_kon': 46, 'pokr_i': 47, 'shir_i': 48, 'ball_i': 49,}
table_2 = context_table(table_2_cells, sheet)

for i in table_2:
    print(isinstance(i['km_nach'], float))

print('----------------------------------------------------')

table_3_cells = {'km': 51, 'ball_i': 52, 'kpr_i': 53, }
table_3 = context_table(table_3_cells, sheet)

for i in table_3:
    print(i)

print('----------------------------------------------------')

table_4_cells = {'km': 56, 'kpr_i': 57, 'E_i': 58, }
table_4 = context_table(table_4_cells, sheet)

for i in table_4:
    print(i)


print(workbook['У 25']['K2'].value)

table_cells_fed = {'num': 1, 'nazvanie': 2, 'cat': 3, 'pokr': 4, 'nagr': 5, 'protyazh': 6, 'prinad': 7, }

def context_start_table(znachenie, sheet, table_cells):
    table = []
    for i in range(2, len(sheet['A'])):
        if sheet.cell(row=i, column=8).value == znachenie:
            table.append({key: sheet.cell(i, table_cells[key]).value for key in table_cells})
    return table

table_fed = context_start_table('федеральная', workbook['В обсл'], table_cells_fed)

def change_table_0(table):
    for row in table:
        row['nagr'] = format_str_or_digit_value(row['nagr'], 1)
        row['protyazh'] = format_int_value(row['protyazh'])
        

for i in table_fed:
    print(i)

change_table_0(table_fed)

for i in table_fed:
    print(i)


