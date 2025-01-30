import openpyxl as xl
from docxtpl import DocxTemplate


workbook = xl.load_workbook('Ведомость.xlsx', data_only=True)
def wgs_table(workbook):

    
    sheet = workbook['В обсл']

    table_cells_fed = {'num_f': 1, 'nazvanie_f': 2, 'protyazh_f': 6,
                       'sh_nach_f': 9, 'dol_nach_f': 10, 'sh_kon_f': 11, 'dol_kon_f': 12, }
    table_cells_reg = {'num_r': 1, 'nazvanie_r': 2, 'protyazh_r': 6,
                       'sh_nach_r': 9, 'dol_nach_r': 10, 'sh_kon_r': 11, 'dol_kon_r': 12, }
    table_cells_mest = {'num_m': 1, 'nazvanie_m': 2, 'protyazh_m': 6,
                        'sh_nach_m': 9, 'dol_nach_m': 10, 'sh_kon_m': 11, 'dol_kon_m': 12, }
    table_cells_chas = {'num_c': 1, 'nazvanie_c': 2, 'protyazh_c': 6,
                        'sh_nach_c': 9, 'dol_nach_c': 10, 'sh_kon_c': 11, 'dol_kon_c': 12, }
    table_cells_les = {'num_l': 1, 'nazvanie_l': 2, 'protyazh_l': 6,
                       'sh_nach_l': 9, 'dol_nach_l': 10, 'sh_kon_l': 11, 'dol_kon_l': 12, }
    table_cells_ved = {'num_v': 1, 'nazvanie_v': 2, 'protyazh_v': 6,
                       'sh_nach_v': 9, 'dol_nach_v': 10, 'sh_kon_v': 11, 'dol_kon_v': 12, }


    # def add_null(x):
    #     if x.isdigit():
    #         return x + ',0'
    #     else:
    #         return x.replace('.', ',')


    def context_table(znachenie, sheet, table_cells):
        """Делает список словарей для таблиц по ключам и номерам столбцов из table_cells для листа sheet"""
        table = []
        for i in range(2, len(sheet['A'])):
            if sheet.cell(row=i, column=8).value == znachenie:
                table.append({key: str(sheet.cell(i, table_cells[key]).value) for key in table_cells})
        # print(table)
        return table


    # print(context_table('региональная', sheet))

    table_fed = context_table('федеральная', sheet, table_cells_fed)
    table_reg = context_table('региональная', sheet, table_cells_reg)
    table_mestn = context_table('местная', sheet, table_cells_mest)
    table_chas = context_table('частная', sheet, table_cells_chas)
    table_les = context_table('лесная', sheet, table_cells_les)
    table_ved = context_table('ведомственная', sheet, table_cells_ved)




    def change_table_0(table):
        for i in table:
            for k in i.keys():
                if i[k].isdigit():
                    i[k] += ',0'
                elif 'protyazh' in k:
                    i[k] = i[k].replace('.', ',')


    change_table_0(table_fed)
    change_table_0(table_reg)
    change_table_0(table_mestn)
    change_table_0(table_chas)
    change_table_0(table_les)
    change_table_0(table_ved)


    context = {
            'table_fed': table_fed,
            'table_reg': table_reg,
            'table_mestn': table_mestn,
            'table_chas': table_chas,
            'table_les': table_les,
            'table_ved': table_ved,
            }


    template = DocxTemplate('templates/Шаблон_таблица_WGS.docx')
    template.render(context)
    template.save(f'Таблица_WGS.docx')


# wgs_table(workbook)
