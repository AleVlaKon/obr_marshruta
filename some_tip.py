import openpyxl as xl
from services import *




# workbook = xl.load_workbook('Ведомость.xlsx', data_only=True)
# sheet_names = workbook.sheetnames
# sheet_1 = workbook['ПК 15']

def some_tips(sheet_1):

    def corr_shir(sheet_1, i):         
        # Если ширина вида 11,5х2 переводит ее в число 
        if type(sheet_1.cell(row = i, column=4).value) in (int, float):
            return round(sheet_1.cell(row = i, column=4).value, 1)
        else:
            razdel_shir = sheet_1.cell(row = i, column=4).value.split('х')
            # print(razdel_shir)
            razdel_shir[0] = razdel_shir[0].replace(',', '.')
            # print(razdel_shir)
            # print(round(float(razdel_shir[0]) * int(razdel_shir[1]), 1))
            return round(float(razdel_shir[0]) * int(razdel_shir[1]), 1)

    perechen_pokritiy = []


    for i in range(12, len(sheet_1['A'])):
        cell_i = sheet_1.cell(row=i, column=5).value
        if cell_i not in [None, 'None'] and cell_i not in perechen_pokritiy:
            perechen_pokritiy.append(sheet_1.cell(row = i, column=5).value)


    pokritia_prot = {i: 0 for i in perechen_pokritiy}
    shirina_min = {i: 100 for i in perechen_pokritiy}
    shirina_max = {i: 0 for i in perechen_pokritiy}


    for i in range(12, len(sheet_1['A'])):
        if sheet_1.cell(row=i, column=1).value not in [None, 'None']:
            pokr = sheet_1.cell(row = i, column=5).value
            prot = round(sheet_1.cell(row = i, column=3).value, 3)
            # shir = round(sheet_1.cell(row = i, column=4).value, 1)
            shir = corr_shir(sheet_1, i)
            pokritia_prot[pokr] += prot
            shirina_min[pokr] = round(shir, 3) if shir < shirina_min[pokr] else shirina_min[pokr]
            shirina_max[pokr] = round(shir, 3) if shir > shirina_max[pokr] else shirina_max[pokr]


    def corr_dict(pokritia):
        for key in pokritia:
            pokritia[key] = format_int_value(pokritia[key])
            # if type(pokritia[key]) is int:
            #     pokritia[key] = f'{pokritia[key]},0'
            # else:
            #     pokritia[key] = str(pokritia[key]).replace('.', ',')

    corr_dict(pokritia_prot)
    corr_dict(shirina_min)
    corr_dict(shirina_max)

    # print(pokritia_prot)
    # print(shirina_min)
    # print(shirina_max)


    opisanie_tipa = {
        1: 'нежесткая, облегченного типа с покрытием из асфальтобетона', 
        2: 'низшего типа с покрытием из песчано-гравийной смеси', 
        3: 'жесткая, капитального типа с покрытием из железобетонных плит', 
        4: 'нежесткая, переходного типа с щебеночным покрытием', 
        5: 'нежесткая, переходного типа с гравийным покрытием', 
        6: 'грунтовая дорога', 
        7: 'грунтощебень',
        8: 'жесткая, капитального типа с бетонным покрытием',
        }


    vid_pokr = {
        'asf': ('асфальтобетон', 'асфальтобетонное'),
        'PGS': ('ПГС',),
        'plit': ('плиты',),
        'sheb': ('щебеночное', 'щебень'),
        'grav': ('гравийное', 'гравий'),
        'grunt': ('грунт', 'грунтовая', 'грунтовое'),
        'grsheb': ('грунтощебень',),
        'beton': ('цементобетон', 'цементобетонное', 'бетонное')
    }



    def add_str(str_i):
        for i in pokritia_prot.keys():
            shir_ob = f'{shirina_min[i]}' if shirina_min[i] == shirina_max[i] else f'от {shirina_min[i]} до {shirina_max[i]}'
            if i in vid_pokr['asf']:
                str_i += f'- {opisanie_tipa[1]}, шириной {shir_ob} м, протяженностью {pokritia_prot[i]} км;\n'
            elif i in vid_pokr['PGS']:
                str_i += f'- {opisanie_tipa[2]}, шириной {shir_ob} м, протяженностью {pokritia_prot[i]} км;\n'
            elif i in vid_pokr['plit']:
                str_i += f'- {opisanie_tipa[3]}, шириной {shir_ob} м, протяженностью {pokritia_prot[i]} км;\n'
            elif i in vid_pokr['sheb']:
                str_i += f'- {opisanie_tipa[4]}, шириной {shir_ob} м, протяженностью {pokritia_prot[i]} км;\n'            
            elif i in vid_pokr['grav']:
                str_i += f'- {opisanie_tipa[5]}, шириной {shir_ob} м, протяженностью {pokritia_prot[i]} км;\n'
            elif i in vid_pokr['grunt']:
                str_i += f'- {opisanie_tipa[6]}, шириной {shir_ob} м, протяженностью {pokritia_prot[i]} км;\n'            
            elif i in vid_pokr['grsheb']:
                str_i += f'- {opisanie_tipa[7]}, шириной {shir_ob} м, протяженностью {pokritia_prot[i]} км;\n'      
            elif i in vid_pokr['beton']:
                str_i += f'- {opisanie_tipa[8]}, шириной {shir_ob} м, протяженностью {pokritia_prot[i]} км;\n' 
            else:
                str_i += f'неизвестный тип покрытия - {i}'
        return str_i

    if len(pokritia_prot) == 1:
        str_i = 'Конструкция дорожной одежды на всём протяжении маршрута'
        str_i = add_str(str_i)
    if len(pokritia_prot) == 2:
        str_i = 'Конструкция дорожной одежды двух типов:\n'
        str_i = add_str(str_i)
    elif len(pokritia_prot) == 3:
        str_i = 'Конструкция дорожной одежды трех типов:\n'
        str_i = add_str(str_i)
    elif len(pokritia_prot) == 4:
        str_i = 'Конструкция дорожной одежды четырех типов:\n'
        str_i = add_str(str_i)

    return str_i.rstrip()

if __name__ == '__main__':
    workbook = xl.load_workbook('Ведомость.xlsx', data_only=True)
    sheet_names = workbook.sheetnames
    sheet_1 = workbook['Ф 3']
    print(some_tips(sheet_1))



