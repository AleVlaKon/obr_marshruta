import openpyxl as xl
from docxtpl import DocxTemplate


workbook = xl.load_workbook('Ведомость тест.xlsx', data_only=True)
# sheet = workbook['В обсл']


def start_table(workbook):
    
    # sheet_names = workbook.sheetnames
    sheet = workbook['В обсл']

    table_cells_fed = {'num_f': 1, 'nazvanie_f': 2, 'cat_f': 3, 'pokr_f': 4, 'nagr_f': 5, 'protyazh_f': 6, 'prinad_f': 7, }
    table_cells_reg = {'num_r': 1, 'nazvanie_r': 2, 'cat_r': 3, 'pokr_r': 4, 'nagr_r': 5, 'protyazh_r': 6, 'prinad_r': 7, }
    table_cells_mest = {'num_m': 1, 'nazvanie_m': 2, 'cat_m': 3, 'pokr_m': 4, 'nagr_m': 5, 'protyazh_m': 6, 'prinad_m': 7, }
    table_cells_chas = {'num_c': 1, 'nazvanie_c': 2, 'cat_c': 3, 'pokr_c': 4, 'nagr_c': 5, 'protyazh_c': 6, 'prinad_c': 7, }
    table_cells_les = {'num_l': 1, 'nazvanie_l': 2, 'cat_l': 3, 'pokr_l': 4, 'nagr_l': 5, 'protyazh_l': 6, 'prinad_l': 7, }
    table_cells_ved = {'num_v': 1, 'nazvanie_v': 2, 'cat_v': 3, 'pokr_v': 4, 'nagr_v': 5, 'protyazh_v': 6, 'prinad_v': 7, }


    def add_null(x):
        ''' Добавляет хвостовой ноль, если х целое число'''
        if x.isdigit():
            return x + ',0'
        else:
            return x.replace('.', ',')
        

    def if_str(value):
        '''Если в ячейке сохранена строка, переводит ее в число'''
        if type(value) == str:
            return float(value.replace(',', '.'))
        else:
            return value


    summas = {'федеральная': 0, 'региональная': 0, 'местная': 0, 'частная': 0, 'лесная': 0, 'ведомственная': 0, }
    for i in range(2, len(sheet['A'])):
        if sheet.cell(row=i, column=8).value in summas:
            summas[sheet.cell(row=i, column=8).value] += if_str(sheet.cell(row=i, column=6).value)
    sum_all = round(sum(summas.values()), 3)     # Протяженность всех дорог
    sum_all = add_null(str(sum_all))             # Протяженность всех дорог (строка)


    for k in summas.keys():
        summas[k] = str(round(summas[k], 3))
        summas[k] = add_null(summas[k])


    def context_table(znachenie, sheet, table_cells):
        # table_cells = {'num': 1, 'nazvanie': 2, 'cat': 3, 'pokr': 4, 'nagr': 5, 'protyazh': 6, 'prinad': 7, }
        """Делает список словарей для таблиц по ключам и номерам столбцов из table_cells для листа sheet"""
        table = []
        for i in range(2, len(sheet['A'])):
            if sheet.cell(row=i, column=8).value == znachenie:
                table.append({key: str(sheet.cell(i, table_cells[key]).value) for key in table_cells})
        # print(table)
        return table


    # print(context_table('региональная', sheet))

    table_fed = context_table('федеральная', sheet, table_cells_fed)
    table_reg = context_table('региональная', sheet, table_cells_reg)
    table_mestn = context_table('местная', sheet, table_cells_mest)
    table_chas = context_table('частная', sheet, table_cells_chas)
    table_les = context_table('лесная', sheet, table_cells_les)
    table_ved = context_table('ведомственная', sheet, table_cells_ved)


    def change_table_0(table):
        '''Заменяет в таблице точки-разделители на запятые 
        и добавляет хвостовой ноль, если значение - целое число'''
        for i in table:
            for k in i.keys():
                if i[k].isdigit():
                    i[k] += ',0'
                elif 'protyazh' in k:
                    i[k] = i[k].replace('.', ',')


    change_table_0(table_fed)
    change_table_0(table_reg)
    change_table_0(table_mestn)
    change_table_0(table_chas)
    change_table_0(table_les)
    change_table_0(table_ved)


    context = {
            'table_fed': table_fed,
            'table_reg': table_reg,
            'table_mestn': table_mestn,
            'table_chas': table_chas,
            'table_les': table_les,
            'table_ved': table_ved,
            'summ_fed': summas['федеральная'],
            'summ_reg': summas['региональная'],
            'summ_mestn': summas['местная'],
            'summ_chas': summas['частная'],
            'summ_les': summas['лесная'],
            'summ_ved': summas['ведомственная'],
            'summ_all': sum_all
            }


    template = DocxTemplate('templates/Шаблон общий.docx')
    template.render(context)
    template.save(f'temp/Шаблон отчета.docx')

# start_table(workbook)
